#!/usr/bin/env python3
"""
CBSE School Scraper — Scrapes CBSE affiliated school data from the SARAS website.
Filters by State and District, exports results to an Excel file.

Usage:
    python scraper.py
    python scraper.py --headless
"""

import sys
import time
import re
import argparse
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
)
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://saras.cbse.gov.in/saras/AffiliatedList/ListOfSchdirReport"

# ───────────────────────────── Browser helpers ─────────────────────────────

def create_driver(headless=False):
    """Create and configure Chrome WebDriver."""
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-gpu")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)
    return driver


def js_click(driver, element):
    """Click via JavaScript (bypasses overlays / scroll issues)."""
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", element)


# ──────────────────────────── User input ───────────────────────────────────

def get_user_input():
    """Prompt for State and District."""
    print("\n" + "=" * 60)
    print("  CBSE School Scraper — SARAS Website")
    print("=" * 60)
    state = input("\nEnter State name (e.g., KARNATAKA): ").strip().upper()
    district = input("Enter District name (e.g., BENGALURU RURAL): ").strip().upper()
    if not state or not district:
        print("Error: Both State and District are required.")
        sys.exit(1)
    return state, district


# ──────────────────────────── Navigation ───────────────────────────────────

def _select_option_by_text(driver, select_id, value):
    """Select an option from a <select> by matching visible text (case-insensitive)."""
    sel = driver.find_element(By.ID, select_id)
    opts = sel.find_elements(By.TAG_NAME, "option")

    # Exact match first
    for opt in opts:
        if opt.text.strip().upper() == value.upper():
            Select(sel).select_by_visible_text(opt.text.strip())
            # Dispatch change event so jQuery/JS handlers fire
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                sel,
            )
            return True

    # Partial match fallback
    for opt in opts:
        if value.upper() in opt.text.strip().upper():
            Select(sel).select_by_visible_text(opt.text.strip())
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                sel,
            )
            return True

    return False


def _list_dropdown_options(driver, select_id):
    """Print all available options for a dropdown (for error messages)."""
    try:
        sel = driver.find_element(By.ID, select_id)
        opts = sel.find_elements(By.TAG_NAME, "option")
        for o in opts:
            txt = o.text.strip()
            if txt and txt != "--select--" and txt != "--Select--":
                print(f"  - {txt}")
    except NoSuchElementException:
        print("  (dropdown not found)")


def navigate_and_search(driver, state, district):
    """Open page → State wise radio → State dropdown → District dropdown → SEARCH."""
    print("\nNavigating to SARAS website...")
    driver.get(URL)
    time.sleep(5)

    # 1) Click "State wise" radio button (id="SearchMainRadioState_wise")
    print("Selecting 'State wise' search mode...")
    try:
        radio = driver.find_element(By.ID, "SearchMainRadioState_wise")
        js_click(driver, radio)
    except NoSuchElementException:
        # Fallback: try clicking the label
        try:
            label = driver.find_element(
                By.CSS_SELECTOR, "label[for='SearchMainRadioState_wise']"
            )
            js_click(driver, label)
        except NoSuchElementException:
            print("  Warning: Could not find 'State wise' radio button.")

    time.sleep(2)

    # 2) Select State (id="State")
    print(f"Selecting State: {state}...")
    if not _select_option_by_text(driver, "State", state):
        print(f"Error: Could not find State '{state}' in the dropdown.")
        print("Available states:")
        _list_dropdown_options(driver, "State")
        sys.exit(1)

    # Wait for district dropdown to populate via AJAX
    time.sleep(4)

    # 3) Select District (id="District")
    print(f"Selecting District: {district}...")
    if not _select_option_by_text(driver, "District", district):
        # Retry after a bit more wait
        time.sleep(3)
        if not _select_option_by_text(driver, "District", district):
            print(f"Error: Could not find District '{district}' in the dropdown.")
            print("Available districts:")
            _list_dropdown_options(driver, "District")
            sys.exit(1)

    time.sleep(1)

    # 4) Submit the form (SEARCH button is an <input type="submit">)
    print("Clicking SEARCH...")
    submitted = False

    # Try finding the submit input
    for selector in [
        "input[type='submit'][value='SEARCH']",
        "input[type='submit'][value='Search']",
        "input[type='submit']",
        "button[type='submit']",
    ]:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, selector)
            js_click(driver, btn)
            submitted = True
            break
        except NoSuchElementException:
            continue

    # Fallback: look for any element with SEARCH text
    if not submitted:
        for tag in ["button", "input", "a"]:
            for el in driver.find_elements(By.TAG_NAME, tag):
                txt = (el.text or el.get_attribute("value") or "").strip()
                if "search" in txt.lower():
                    js_click(driver, el)
                    submitted = True
                    break
            if submitted:
                break

    # Last fallback: submit the form via JS
    if not submitted:
        print("  Using JS form submit as fallback...")
        driver.execute_script("document.querySelector('form').submit();")

    print("Waiting for results...")
    time.sleep(6)

    # 5) Maximise entries per page to reduce pagination
    _maximise_page_size(driver)


def _maximise_page_size(driver):
    """Set the DataTable 'Show entries' dropdown to the largest value (100)."""
    try:
        sel = driver.find_element(By.CSS_SELECTOR, "select[name='myTable_length']")
        Select(sel).select_by_value("100")
        print("  Show entries set to: 100")
        time.sleep(3)
    except NoSuchElementException:
        try:
            # Fallback: any datatable length select
            sel = driver.find_element(By.CSS_SELECTOR, ".dataTables_length select")
            opts = sel.find_elements(By.TAG_NAME, "option")
            if opts:
                Select(sel).select_by_value(opts[-1].get_attribute("value"))
                print(f"  Show entries set to: {opts[-1].text.strip()}")
                time.sleep(3)
        except NoSuchElementException:
            pass


# ──────────────────────────── Scraping ─────────────────────────────────────

def _get_total_entries(driver):
    """Parse 'Showing X to Y of Z Entries' to get Z."""
    try:
        info = driver.find_element(By.ID, "myTable_info")
        m = re.search(r"of\s+([\d,]+)\s+", info.text, re.I)
        if m:
            return int(m.group(1).replace(",", ""))
    except NoSuchElementException:
        pass
    return 0


def _parse_row(row):
    """Extract structured data from a single table <tr> element.

    The table has 7 columns:
      [0] S No
      [1] Aff. No & School Code  — contains <b> tags with labeled data
      [2] State & District        — contains <b> tags with labeled data
      [3] Status
      [4] School & Head Name      — contains <b> tags with labeled data
      [5] Address                  — contains <b> tags with labeled data
      [6] Details (View link)      — we skip this
    """
    cells = row.find_elements(By.TAG_NAME, "td")
    if len(cells) < 6:
        return None

    # Helper: get inner text, handling potential empty cells
    def txt(idx):
        try:
            return cells[idx].text.strip()
        except (IndexError, StaleElementReferenceException):
            return ""

    def html(idx):
        try:
            return cells[idx].get_attribute("innerHTML")
        except (IndexError, StaleElementReferenceException):
            return ""

    # S NO
    s_no = txt(0)

    # AFF. NO & SCHOOL CODE (cell 1)
    aff_text = txt(1)
    aff_no = ""
    school_code = ""
    m = re.search(r"Aff\.?\s*No\.?\s*:?\s*(\S+)", aff_text, re.I)
    if m:
        aff_no = m.group(1).strip().rstrip(",")
    m = re.search(r"Sch\.?\s*Code\s*:?\s*(\S+)", aff_text, re.I)
    if m:
        school_code = m.group(1).strip()
    # Fallback: split by newline
    if not aff_no:
        lines = [l.strip() for l in aff_text.split("\n") if l.strip()]
        if len(lines) >= 1:
            aff_no = lines[0]
        if len(lines) >= 2:
            school_code = lines[1]

    # STATE & DISTRICT (cell 2)
    sd_text = txt(2)
    state_val = ""
    district_val = ""
    m = re.search(r"State\s*:\s*(.+?)(?:\n|District|$)", sd_text, re.I)
    if m:
        state_val = m.group(1).strip()
    m = re.search(r"District\s*:\s*(.+)", sd_text, re.I)
    if m:
        district_val = m.group(1).strip()

    # STATUS (cell 3)
    status = txt(3)

    # SCHOOL & HEAD NAME (cell 4)
    sh_text = txt(4)
    school_name = ""
    principal = ""
    m = re.search(r"Name\s*:\s*(.+?)(?:\n|Head|Principal|$)", sh_text, re.I)
    if m:
        school_name = m.group(1).strip()
    m = re.search(r"(?:Head/?Principal|Principal)\s*Name\s*:?\s*(.+)", sh_text, re.I)
    if m:
        principal = m.group(1).strip()

    # ADDRESS (cell 5)
    addr_text = txt(5)
    address = ""
    website = ""
    m = re.search(r"Address\s*:\s*(.+?)(?:\nWebsite|$)", addr_text, re.I | re.S)
    if m:
        address = m.group(1).strip()
    else:
        address = addr_text
    m = re.search(r"Website\s*:\s*(.+)", addr_text, re.I)
    if m:
        website = m.group(1).strip()

    return {
        "S No": s_no,
        "Affiliation No": aff_no,
        "School Code": school_code,
        "State": state_val,
        "District": district_val,
        "Status": status,
        "School Name": school_name,
        "Head/Principal Name": principal,
        "Address": address,
        "Website": website,
    }


def scrape_all_pages(driver):
    """Iterate through all DataTable pages and collect every row."""
    all_data = []
    page = 1
    total = _get_total_entries(driver)
    if total:
        print(f"Total entries: {total}")

    while True:
        print(f"  Scraping page {page}...")
        rows_data = []
        try:
            tbody = driver.find_element(By.CSS_SELECTOR, "#myTable tbody")
            rows = tbody.find_elements(By.TAG_NAME, "tr")
            for row in rows:
                if "No data available" in row.text:
                    continue
                try:
                    parsed = _parse_row(row)
                    if parsed:
                        rows_data.append(parsed)
                except StaleElementReferenceException:
                    continue
        except NoSuchElementException:
            pass

        if not rows_data:
            if page == 1:
                print("  No data found. Please verify your State and District.")
            break

        all_data.extend(rows_data)
        print(f"    → {len(rows_data)} rows  (total so far: {len(all_data)})")

        if total and len(all_data) >= total:
            break

        # Click Next
        try:
            next_btn = driver.find_element(By.ID, "myTable_next")
            classes = next_btn.get_attribute("class") or ""
            if "disabled" in classes:
                break
            js_click(driver, next_btn)
            time.sleep(2)
            page += 1
        except NoSuchElementException:
            break

    return all_data


# ──────────────────────────── Excel export ─────────────────────────────────

HEADERS = [
    "S No",
    "Affiliation No",
    "School Code",
    "State",
    "District",
    "Status",
    "School Name",
    "Head/Principal Name",
    "Address",
    "Website",
]

COL_WIDTHS = {
    "S No": 8,
    "Affiliation No": 16,
    "School Code": 14,
    "State": 18,
    "District": 22,
    "Status": 22,
    "School Name": 42,
    "Head/Principal Name": 28,
    "Address": 55,
    "Website": 38,
}


def save_to_excel(data, state, district):
    """Write scraped data to a formatted .xlsx file."""
    if not data:
        print("No data to save.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "CBSE Schools"

    # Styles
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    # Write header row
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    # Write data rows
    for r, row_data in enumerate(data, 2):
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=r, column=c, value=row_data.get(h, ""))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border

    # Column widths
    for c, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(h, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    filename = f"CBSE_Schools_{state.replace(' ', '_')}_{district.replace(' ', '_')}.xlsx"
    wb.save(filename)
    return filename


# ──────────────────────────── Main ─────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CBSE School Scraper")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    args = parser.parse_args()

    state, district = get_user_input()

    print(f"\nScraping schools for  State: {state}  |  District: {district}")
    print("Launching browser...")

    driver = None
    try:
        driver = create_driver(headless=args.headless)
        navigate_and_search(driver, state, district)
        all_data = scrape_all_pages(driver)

        if all_data:
            filepath = save_to_excel(all_data, state, district)
            print(f"\n{'=' * 60}")
            print(f"  Done!  Schools scraped: {len(all_data)}")
            print(f"  Output file: {filepath}")
            print(f"{'=' * 60}")
        else:
            print("\nNo schools found for the given State and District.")

    except KeyboardInterrupt:
        print("\nInterrupted by user.")
    except Exception as e:
        print(f"\nFatal error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            driver.quit()
            print("Browser closed.")


if __name__ == "__main__":
    main()
